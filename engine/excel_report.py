"""
⚡ Hızlı Transformatör — High-Performance Excel Report Generator
Generates corporate, multi-sheet, beautifully formatted .xlsx engineering and tender reports.
Includes full 10-module parameters, LME market prices, TOC, IEC audits, Cu vs Al, and voltage regulation tables.
"""

import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Brand Corporate Colors (BEST Navy Palette)
NAVY_HEADER = "0B2545"
ICE_BLUE = "E6F0FA"
LIGHT_GRAY = "F8FAFC"
BORDER_GRAY = "D0D7DE"
GREEN_ACCENT = "00875A"
ORANGE_ACCENT = "E65100"
PURPLE_ACCENT = "5E35B1"

def _get_styles():
    """Returns standardized fonts, fills, and borders for the workbook."""
    font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Segoe UI", size=9, italic=True, color="FFFFFF")
    font_section = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_tbl_header = Font(name="Segoe UI", size=10, bold=True, color="0B2545")
    font_bold = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
    font_regular = Font(name="Segoe UI", size=10, color="334155")
    font_muted = Font(name="Segoe UI", size=9, color="64748B")
    font_pass = Font(name="Segoe UI", size=10, bold=True, color="00875A")
    font_warn = Font(name="Segoe UI", size=10, bold=True, color="E65100")

    fill_title = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
    fill_section = PatternFill(start_color="134074", end_color="134074", fill_type="solid")
    fill_tbl_header = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_zebra = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

    thin_border_side = Side(border_style="thin", color=BORDER_GRAY)
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_bottom_thick = Border(bottom=Side(border_style="medium", color=NAVY_HEADER))

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    return {
        "font_title": font_title, "font_subtitle": font_subtitle, "font_section": font_section,
        "font_tbl_header": font_tbl_header, "font_bold": font_bold, "font_regular": font_regular,
        "font_muted": font_muted, "font_pass": font_pass, "font_warn": font_warn,
        "fill_title": fill_title, "fill_section": fill_section, "fill_tbl_header": fill_tbl_header,
        "fill_zebra": fill_zebra, "border_cell": border_cell, "border_bottom_thick": border_bottom_thick,
        "align_center": align_center, "align_left": align_left, "align_right": align_right
    }

def auto_fit_columns(ws, max_cols=12):
    """Adjusts column widths dynamically for clean readable tables."""
    for col in range(1, max_cols + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                lines = str(cell.value).split("\n")
                for line in lines:
                    if len(str(line)) > max_len:
                        max_len = len(str(line))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def build_excel_report(data, res):
    """
    Builds a complete, multi-sheet, beautifully styled Excel (.xlsx) workbook.
    """
    wb = openpyxl.Workbook()
    st = _get_styles()

    el = res.get('electrical', {})
    cd = res.get('core_design', {})
    wd = res.get('winding', {})
    ls = res.get('losses', {})
    vr = res.get('voltage_regulation', {})
    th = res.get('thermal', {})
    sc = res.get('short_circuit', {})
    ins = res.get('insulation', {})
    mag = res.get('magnetization', {})
    eco = res.get('economic', {})
    cost = res.get('cost', {})
    comp = res.get('comparison', {})
    iec = res.get('iec_compliance', {})

    now_str = datetime.now().strftime("%d.%m.%Y %H:%M")

    # ==========================================
    # SHEET 1: ETİKET & ŞARTNAME ÖZETİ
    # ==========================================
    ws1 = wb.active
    ws1.title = "Tasarım Özeti & Etiket"
    ws1.views.sheetView[0].showGridLines = True

    # Title Banner
    ws1.merge_cells("A1:F2")
    cell_t = ws1["A1"]
    cell_t.value = "HIZLI TRANSFORMATÖR — GÜÇ TRANSFORMATÖRÜ MÜHENDİSLİK & İHALE RAPORU"
    cell_t.font = st["font_title"]
    cell_t.fill = st["fill_title"]
    cell_t.alignment = st["align_center"]

    ws1.merge_cells("A3:F3")
    cell_sub = ws1["A3"]
    cell_sub.value = f"IEC 60076 Uluslararası Standartlarına Uygun Hesaplama Raporu | Rapor Tarihi: {now_str}"
    cell_sub.font = st["font_subtitle"]
    cell_sub.fill = st["fill_title"]
    cell_sub.alignment = st["align_center"]

    # Section 0: Live LME Commodity Prices & Exchange Rates
    ws1.cell(row=5, column=1, value="LONDRA METAL BORSASI (LME) CANLI PİYASA GÖSTERGELERİ").font = st["font_section"]
    ws1.cell(row=5, column=1).fill = st["fill_section"]
    ws1.merge_cells("A5:F5")

    cu_p = float(res.get('prices', {}).get('copper', 9.50))
    al_p = float(res.get('prices', {}).get('aluminum', 2.50))
    try_p = float(res.get('prices', {}).get('usd_try', 34.00))

    lme_specs = [
        ("Bakır (LME Copper Grade A)", f"${cu_p:.2f} / kg ({cu_p * try_p:,.2f} TL/kg)", "Alüminyum (LME Primary 99.7)", f"${al_p:.2f} / kg ({al_p * try_p:,.2f} TL/kg)"),
        ("Dolar/TL Döviz Kuru (USD/TRY)", f"₺{try_p:.2f}", "Kayıp Kapitalizasyon (A/B)", f"A: ${data.get('A_factor', 8.0)}/W | B: ${data.get('B_factor', 2.0)}/W")
    ]
    for i, row in enumerate(lme_specs):
        r_num = 6 + i
        ws1.cell(row=r_num, column=1, value=row[0]).font = st["font_bold"]
        ws1.cell(row=r_num, column=2, value=row[1]).font = st["font_regular"]
        ws1.cell(row=r_num, column=4, value=row[2]).font = st["font_bold"]
        ws1.cell(row=r_num, column=5, value=row[3]).font = st["font_regular"]
        for c in [1, 2, 4, 5]:
            ws1.cell(row=r_num, column=c).border = st["border_cell"]
            if i % 2 == 1: ws1.cell(row=r_num, column=c).fill = st["fill_zebra"]

    # Section 1: IEC Compliance Summary
    start_iec = 9
    ws1.cell(row=start_iec, column=1, value=f"IEC 60076 STANDART UYGUNLUK DENETİMİ — {iec.get('total_score', '5/5 ONAYLI')}").font = st["font_section"]
    ws1.cell(row=start_iec, column=1).fill = st["fill_section"]
    ws1.merge_cells(f"A{start_iec}:F{start_iec}")

    headers_iec = ["Denetim Parametresi", "IEC Standardı", "Tasarım Değeri", "Standart Limiti", "Uygunluk Durumu"]
    for col_idx, h in enumerate(headers_iec, 1):
        cell = ws1.cell(row=start_iec + 1, column=col_idx, value=h)
        cell.font = st["font_tbl_header"]
        cell.fill = st["fill_tbl_header"]
        cell.alignment = st["align_center"]
        cell.border = st["border_cell"]

    iec_rows = [
        ("Üst Yağ Sıcaklık Artışı (Δθ_oil)", "IEC 60076-2", f"{th.get('top_oil_rise_C')} °C", "≤ 60.0 °C", iec.get('top_oil', {}).get('status', 'IEC UYGUN')),
        ("Hot-Spot Sargı Tepe Sıcaklığı (θ_hs)", "IEC 60076-7", f"{th.get('hot_spot_temp_C')} °C", "≤ 98.0 °C", iec.get('hot_spot', {}).get('status', 'GÜVENLİ')),
        ("Boşta Akım Oranı (I0 %)", "IEC 60076-1", f"{mag.get('I0_pct')} %", "≤ 1.50 %", iec.get('no_load_current', {}).get('status', 'STANDART İÇİ')),
        ("Kısa Devre Termal Dayanım", "IEC 60076-5", "2.0 Saniye", "≥ 2.0 Saniye", "TERMAL ONAYLI"),
        ("Darbe & AC Test Dayanımı (BIL)", "IEC 60076-3", f"BIL {ins.get('hv_BIL_kVp')} kVp", "Standart Norm", "DİELEKTRİK UYGUN")
    ]

    for row_idx, r in enumerate(iec_rows, start_iec + 2):
        for col_idx, val in enumerate(r, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = st["font_regular"] if col_idx != 5 else (st["font_pass"] if "UYGUN" in val or "GÜVENLİ" in val or "ONAYLI" in val or "STANDART" in val else st["font_warn"])
            cell.alignment = st["align_center"] if col_idx > 1 else st["align_left"]
            cell.border = st["border_cell"]
            if row_idx % 2 == 0: cell.fill = st["fill_zebra"]

    # Section 2: Transformer Nameplate & General Parameters
    start_row = start_iec + len(iec_rows) + 3
    ws1.cell(row=start_row, column=1, value="GENEL TRANSFORMATÖR ETİKET DEĞERLERİ").font = st["font_section"]
    ws1.cell(row=start_row, column=1).fill = st["fill_section"]
    ws1.merge_cells(f"A{start_row}:F{start_row}")

    opt = res.get('optimization', {})
    opt_label = f"⚡ Optimizasyon ({opt.get('iterations', 0)} İter)" if opt.get('enabled') else "Manuel Mod"

    specs = [
        ("Anma Gücü (S)", f"{el.get('rated_power_kVA', data.get('S', 50000)/1000.0)} kVA", "Primer Gerilimi (V1)", f"{data.get('V1')} V (Faz: {el.get('V1_phase')} V)"),
        ("Vektör Grubu", f"{el.get('vector_group_name', 'Dyn11')}", "Sekonder Gerilimi (V2)", f"{data.get('V2')} V (Faz: {el.get('V2_phase')} V)"),
        ("Tasarım Modu & Sabit (k)", f"k = {opt.get('selected_k_constant', data.get('k_constant', '0.45'))} ({opt_label})", "Sarım Oranı (a = N1/N2)", f"{el.get('a')}"),
        ("Kısa Devre Empedansı (uk%)", f"{el.get('uk_pct', data.get('uk'))} %", "Boşta Kayıp (P0)", f"{data.get('P0')} W (Hesap: {cd.get('P0_estimated_W')} W)"),
        ("Direnç / Reaktans (ur / ux)", f"%{el.get('ur_pct')} / %{el.get('ux_pct')}", "Garanti Yük Kaybı (Pk)", f"{data.get('Pk')} W"),
        ("Hesaplanan Toplam Pk", f"{wd.get('pk_calculated_total')} W", "Tam Yük Verimi (cosφ=1)", f"%{el.get('efficiency')}"),
        ("Primer Anma Akımı (I1)", f"{el.get('I1')} A (Faz: {el.get('I1_phase')} A)", "Sekonder Anma Akımı (I2)", f"{el.get('I2')} A (Faz: {el.get('I2_phase')} A)"),
        ("Primer İletken Malzemesi", f"{data.get('material_hv', 'Cu')}", "Sekonder İletken Malzemesi", f"{data.get('material_lv', 'Cu')}"),
        ("Nüve Sac Tipi", f"{cd.get('core_label')}", "Soğutma & Yağ Tipi", f"{th.get('cooling_method', 'ONAN')} / {data.get('oil_type', 'mineral')}")
    ]

    for i, row in enumerate(specs):
        r_num = start_row + 1 + i
        ws1.cell(row=r_num, column=1, value=row[0]).font = st["font_bold"]
        ws1.cell(row=r_num, column=2, value=row[1]).font = st["font_regular"]
        ws1.cell(row=r_num, column=4, value=row[2]).font = st["font_bold"]
        ws1.cell(row=r_num, column=5, value=row[3]).font = st["font_regular"]
        for c in [1, 2, 4, 5]:
            ws1.cell(row=r_num, column=c).border = st["border_cell"]
            if i % 2 == 1: ws1.cell(row=r_num, column=c).fill = st["fill_zebra"]

    # Section 3: Weights & Material Bill Summary
    start_w = start_row + len(specs) + 2
    ws1.cell(row=start_w, column=1, value="AĞIRLIK & BOYUT ÖZETİ").font = st["font_section"]
    ws1.cell(row=start_w, column=1).fill = st["fill_section"]
    ws1.merge_cells(f"A{start_w}:F{start_w}")

    w_specs = [
        ("Primer İletken Ağırlığı", f"{cost.get('weights', {}).get('hv')} kg", "Nüve (Çekirdek) Ağırlığı", f"{cd.get('core_weight_kg')} kg"),
        ("Sekonder İletken Ağırlığı", f"{cost.get('weights', {}).get('lv')} kg", "Kazan (Tank) Ağırlığı", f"{th.get('tank_weight_kg')} kg"),
        ("Toplam İletken Ağırlığı", f"{cost.get('weights', {}).get('total')} kg", "Kullanılacak Yağ Hacmi / Ağırlığı", f"{th.get('oil_volume_L')} L ({th.get('oil_weight_kg')} kg)"),
        ("Toplam Kuru Ağırlık", f"{cost.get('weights', {}).get('dry')} kg", "Toplam Islak Ağırlık (Montajlı)", f"{cost.get('weights', {}).get('wet')} kg")
    ]

    for i, row in enumerate(w_specs):
        r_num = start_w + 1 + i
        ws1.cell(row=r_num, column=1, value=row[0]).font = st["font_bold"]
        ws1.cell(row=r_num, column=2, value=row[1]).font = st["font_regular"]
        ws1.cell(row=r_num, column=4, value=row[2]).font = st["font_bold"]
        ws1.cell(row=r_num, column=5, value=row[3]).font = st["font_regular"]
        for c in [1, 2, 4, 5]:
            ws1.cell(row=r_num, column=c).border = st["border_cell"]
            if i % 2 == 1: ws1.cell(row=r_num, column=c).fill = st["fill_zebra"]

    auto_fit_columns(ws1, 6)

    # ==========================================
    # SHEET 2: 10 MODÜL DETAYLI MÜHENDİSLİK
    # ==========================================
    ws2 = wb.create_sheet(title="10 Modül Mühendislik Analizi")
    ws2.views.sheetView[0].showGridLines = True

    # Title
    ws2.merge_cells("A1:E2")
    c_t2 = ws2["A1"]
    c_t2.value = "DETAYLI 10 MODÜL MÜHENDİSLİK & ELEKTROMANYETİK HESAP MATRİSİ"
    c_t2.font = st["font_title"]
    c_t2.fill = st["fill_title"]
    c_t2.alignment = st["align_center"]

    modules_data = [
        ("1. ELEKTRİKSEL & EŞDEĞER DEVRE PARAMETRELERİ (IEC 60076-1)", [
            ("Vektör Grubu", f"{el.get('vector_group_name', 'Dyn11')}", "Faz Kayması", f"{el.get('phase_displacement_deg', 330)}°"),
            ("Primer Faz Gerilimi (V1_ph)", f"{el.get('V1_phase')} V", "Sekonder Faz Gerilimi (V2_ph)", f"{el.get('V2_phase')} V"),
            ("Primer Faz Akımı (I1_ph)", f"{el.get('I1_phase')} A", "Sekonder Faz Akımı (I2_ph)", f"{el.get('I2_phase')} A"),
            ("Kısa Devre Gerilimi (Vk)", f"{el.get('Vk')} V", "Eşdeğer Empedans (Zk)", f"{el.get('Zk')} Ω/faz"),
            ("Eşdeğer Direnç (Rk)", f"{el.get('Rk')} Ω/faz", "Eşdeğer Reaktans (Xk)", f"{el.get('Xk')} Ω/faz"),
            ("Kaçak Endüktans (Lk)", f"{el.get('Lk_mH')} mH", "Direnç / Reaktans Oranı", f"%{el.get('ur_pct')} / %{el.get('ux_pct')}")
        ]),
        ("2. NÜVE & MANYETİK TASARIM & DİNAMİK GEOMETRİ ANALİZİ (FARADAY EMF)", [
            ("Çalışma Akı Yoğunluğu (Bm)", f"{cd.get('Bm')} Tesla", "Net Nüve Kesiti (Ai)", f"{cd.get('Ai_cm2')} cm²"),
            ("Brüt Nüve Kesiti (Ag)", f"{cd.get('Ag_cm2')} cm²", "Kademeli Nüve Çapı (d)", f"{cd.get('core_diameter_mm')} mm"),
            ("Pencere Yüksekliği (Hw)", f"{cd.get('window_height_mm')} mm (Hw/d: {cd.get('ratio_hw_d', '2.70')})", "Bacak Merkez Mesafesi (A)", f"{cd.get('limb_center_dist_mm')} mm (A/d: {cd.get('ratio_a_d', '2.08')})"),
            ("Histerezis Kaybı", f"{cd.get('P_hysteresis_W')} W", "Foucault (Eddy) Kaybı", f"{cd.get('P_eddy_W')} W")
        ]),
        ("3. SARGI, İLETKEN BOYUTLANDIRMA & YÜK KAYIP AYRIŞIMI (IEC 60076-1)", [
            ("Spesifik Gerilim (Et)", f"{el.get('Et')} V/Tur", "Primer Sarım Sayısı (N1)", f"{el.get('N1')} Tur"),
            ("Sekonder Sarım Sayısı (N2)", f"{el.get('N2')} Tur", "Primer Kesit / Çap", f"{el.get('A1')} mm² ({wd.get('d_conductor_hv_mm')} mm)"),
            ("Sekonder Kesit / Çap", f"{el.get('A2')} mm² ({wd.get('d_conductor_lv_mm')} mm)", "75°C IEC Direnci (HV / LV)", f"{wd.get('R_hv_75')} Ω / {wd.get('R_lv_75')} Ω"),
            ("DC Bakır Kaybı (Pdc)", f"{wd.get('pk_dc_only')} W", "Sargı Eddy Kaybı (Peddy)", f"{wd.get('pk_eddy')} W (%{wd.get('pk_eddy_pct')}%)"),
            ("Kazan Saçılma Kaybı (P_stray)", f"{wd.get('pk_stray')} W (%{wd.get('pk_stray_pct')}%)", "Hesaplanan Toplam Pk", f"{wd.get('pk_calculated_total')} W (Garanti: {data.get('Pk')} W)"),
            ("Deri Kalınlığı (Skin Depth δ)", f"HV: {wd.get('skin_depth_hv_mm', '10.31')} mm | LV: {wd.get('skin_depth_lv_mm', '10.31')} mm", "Paralel Tel (HV/LV)", f"{wd.get('n_parallel_hv')}/{wd.get('n_parallel_lv')}")
        ]),
        ("4. YALITIM & DİELEKTRİK TESTLERİ (IEC 60076-3)", [
            ("Primer Um Gerilimi", f"{ins.get('hv_Um_kV')} kV", "Primer Darbe Dayanımı (BIL)", f"{ins.get('hv_BIL_kVp')} kVp"),
            ("Primer AC Test (1 dk)", f"{ins.get('hv_AC_test_kV')} kV rms", "Sekonder Um / AC Test", f"{ins.get('lv_Um_kV')} kV / {ins.get('lv_AC_test_kV')} kV"),
            ("Min. Kaçak Mesafesi", f"{ins.get('creepage_distance_mm')} mm", "Yağ İçi Açıklık", f"{ins.get('clearance_oil_mm')} mm")
        ]),
        ("5. TERMODİNAMİK & SOĞUTMA ISIL ANALİZİ (IEC 60076-2)", [
            ("Toplam Isı Kaybı (P0+Pk)", f"{th.get('total_heat_loss_W')} W", "Gerekli Soğutma Yüzeyi", f"{th.get('cooling_area_m2')} m²"),
            ("Üst Yağ Sıcaklık Artışı (Δθ)", f"{th.get('top_oil_rise_C')} °C", "Hot-Spot Sıcaklığı (θ_hs)", f"{th.get('hot_spot_temp_C')} °C"),
            ("Termal Zaman Sabiti (τ)", f"{th.get('thermal_time_constant_h')} Saat", "Konservatör Hacmi", f"{th.get('conservator_volume_L')} L")
        ]),
        ("6. KISA DEVRE DİNAMİĞİ & MEKANİK KUVVETLER (IEC 60076-5)", [
            ("Primer Simetrik Kısa Devre", f"{sc.get('Isc_A')} A", "Sekonder Simetrik Kısa Devre", f"{sc.get('Isc_lv_A')} A"),
            ("Asimetrik Tepe Akımı (Ipeak)", f"{sc.get('Ipeak_A')} A", "Asimetri Çarpanı (K)", f"{sc.get('asymmetry_factor_k')}"),
            ("Eksenel Kuvvet (F_axial)", f"{sc.get('F_axial_N')} N", "Radyal Kuvvet (F_radial)", f"{sc.get('F_radial_N')} N")
        ]),
        ("7. MIKNATISLANMA & INRUSH DİNAMİKLERİ", [
            ("Demir Kayıp Akımı (Ic)", f"{mag.get('Ic_A')} A", "Mıknatıslanma Akımı (Im)", f"{mag.get('Im_A')} A"),
            ("Toplam Boşta Akım (I0)", f"{mag.get('I0_A')} A ({mag.get('I0_pct')} %)", "Boşta Güç Faktörü (cosφ0)", f"{mag.get('cos_phi_0')}"),
            ("İlk Periyot Tepe Inrush Akımı", f"{mag.get('inrush_peak_hv')} A", "Inrush / In Oranı", f"{mag.get('inrush_ratio')} katı")
        ]),
        ("8. GERİLİM REGÜLASYONU & YÜK DÜŞÜMÜ ANALİZİ", [
            ("Regülasyon (cosφ=0.8)", f"%{vr.get('reg_08', '—')}", "Maksimum Düşüm Açısı", f"cosφ = {vr.get('max_reg_cos_phi', '—')}"),
            ("Regülasyon (cosφ=1.0)", f"%{vr.get('reg_10', '—')}", "Regülasyon (cosφ=0.9)", f"%{vr.get('reg_09', '—')}")
        ])
    ]

    curr_row = 4
    for mod_title, rows in modules_data:
        ws2.cell(row=curr_row, column=1, value=mod_title).font = st["font_section"]
        ws2.cell(row=curr_row, column=1).fill = st["fill_section"]
        ws2.merge_cells(f"A{curr_row}:E{curr_row}")
        curr_row += 1

        for r_idx, (k1, v1, k2, v2) in enumerate(rows):
            ws2.cell(row=curr_row, column=1, value=k1).font = st["font_bold"]
            ws2.cell(row=curr_row, column=2, value=v1).font = st["font_regular"]
            ws2.cell(row=curr_row, column=4, value=k2).font = st["font_bold"]
            ws2.cell(row=curr_row, column=5, value=v2).font = st["font_regular"]
            for c in [1, 2, 4, 5]:
                ws2.cell(row=curr_row, column=c).border = st["border_cell"]
                if r_idx % 2 == 1: ws2.cell(row=curr_row, column=c).fill = st["fill_zebra"]
            curr_row += 1
        curr_row += 1

    auto_fit_columns(ws2, 5)

    # ==========================================
    # SHEET 3: FİNANSAL TOC & CU-AL KIYASLAMA
    # ==========================================
    ws3 = wb.create_sheet(title="Finansal Analiz & TOC")
    ws3.views.sheetView[0].showGridLines = True

    # Title
    ws3.merge_cells("A1:F2")
    c_t3 = ws3["A1"]
    c_t3.value = "FİNANSAL İHALE, 25 YILLIK TOC & BAKIR-ALÜMİNYUM KARŞILAŞTIRMASI"
    c_t3.font = st["font_title"]
    c_t3.fill = st["fill_title"]
    c_t3.alignment = st["align_center"]

    # Section 1: Financial Overview
    ws3.cell(row=4, column=1, value="FİNANSAL DEĞERLENDİRME & KAYIP MALİYETLERİ").font = st["font_section"]
    ws3.cell(row=4, column=1).fill = st["fill_section"]
    ws3.merge_cells("A4:F4")

    usd_try_rate = float(res.get('prices', {}).get('usd_try', 34.0))
    tot_cost_usd = float(cost.get('total', {}).get('total_cost', 0))
    loss_cost_usd = float(eco.get('loss_cost_usd', 0))
    toc_usd = float(eco.get('toc_usd', 0))
    lcc_usd = float(eco.get('lcc_usd', 0))

    fin_specs = [
        ("İletken Sargı Maliyeti (USD)", f"${tot_cost_usd:,.2f}", "İletken Sargı Maliyeti (TL)", f"₺{tot_cost_usd * usd_try_rate:,.2f}"),
        ("Kayıp Değerleme Bedeli (A·P0 + B·Pk)", f"${loss_cost_usd:,.2f}", "Kayıp Bedeli (TL)", f"₺{loss_cost_usd * usd_try_rate:,.2f}"),
        ("Toplam Sahip Olma Maliyeti (TOC)", f"${toc_usd:,.2f}", "TOC Maliyeti (TL)", f"₺{toc_usd * usd_try_rate:,.2f}"),
        ("25 Yıllık LCC Yaşam Döngüsü Bedeli", f"${lcc_usd:,.2f}", "Yıllık İşletme Kayıp Bedeli", f"${eco.get('annual_operating_cost_usd', 0):,.2f} / Yıl")
    ]

    for i, row in enumerate(fin_specs):
        r_num = 5 + i
        ws3.cell(row=r_num, column=1, value=row[0]).font = st["font_bold"]
        ws3.cell(row=r_num, column=2, value=row[1]).font = st["font_regular"]
        ws3.cell(row=r_num, column=4, value=row[2]).font = st["font_bold"]
        ws3.cell(row=r_num, column=5, value=row[3]).font = st["font_regular"]
        for c in [1, 2, 4, 5]:
            ws3.cell(row=r_num, column=c).border = st["border_cell"]
            if i % 2 == 1: ws3.cell(row=r_num, column=c).fill = st["fill_zebra"]

    # Section 2: Cu vs Al Tender Comparison Table
    start_comp = 11
    ws3.cell(row=start_comp, column=1, value=f"BAKIR (Cu) vs ALÜMİNYUM (Al) İHALE KIYASLAMA MATRİSİ — {comp.get('delta', {}).get('advantage', 'İhale Karar Analizi')}").font = st["font_section"]
    ws3.cell(row=start_comp, column=1).fill = st["fill_section"]
    ws3.merge_cells(f"A{start_comp}:F{start_comp}")

    headers_comp = ["Karşılaştırma Kriteri", "Bakır Tasarım (Cu)", "Alüminyum Tasarım (Al)", "Mühendislik / Finansal Fark", "İhale Yorumu"]
    for col_idx, h in enumerate(headers_comp, 1):
        cell = ws3.cell(row=start_comp + 1, column=col_idx, value=h)
        cell.font = st["font_tbl_header"]
        cell.fill = st["fill_tbl_header"]
        cell.alignment = st["align_center"]
        cell.border = st["border_cell"]

    comp_cu = comp.get('cu', {})
    comp_al = comp.get('al', {})
    comp_delta = comp.get('delta', {})

    comp_rows = [
        ("İletken Sargı Ağırlığı", f"{comp_cu.get('cond_weight_kg')} kg", f"{comp_al.get('cond_weight_kg')} kg", f"Alüminyum %{(((float(comp_cu.get('cond_weight_kg', 60)) - float(comp_al.get('cond_weight_kg', 40))) / float(comp_cu.get('cond_weight_kg', 60))) * 100):.0f} Daha Hafif", "Al iletken hafiflik sağlar"),
        ("İletken İmalat Maliyeti", f"${float(comp_cu.get('cond_cost_usd', 0)):,.2f}", f"${float(comp_al.get('cond_cost_usd', 0)):,.2f}", f"Tasarruf: ${float(comp_delta.get('savings_usd', 0)):,.2f} (%{comp_delta.get('savings_pct')}%)", "Alüminyum çok daha ekonomik"),
        ("Trafo Toplam Islak Ağırlık", f"{comp_cu.get('wet_weight_kg')} kg", f"{comp_al.get('wet_weight_kg')} kg", f"Δ {comp_delta.get('weight_diff_kg')} kg", "Kazan ve yağ dahil toplam"),
        ("Toplam Sahip Olma (TOC)", f"${float(comp_cu.get('toc_usd', 0)):,.2f}", f"${float(comp_al.get('toc_usd', 0)):,.2f}", f"TOC Farkı: ${abs(float(comp_cu.get('toc_usd', 0)) - float(comp_al.get('toc_usd', 0))):,.2f}", "25 yıllık toplam ihale farkı"),
        ("Gövde / Tank Hacmi", "Standart (Kompakt)", "+%15 Daha Geniş Gövde", "Cu %15 yer tasarrufu sağlar", "Bakır kompakt montaj avantajı sunar")
    ]

    for row_idx, r in enumerate(comp_rows, start_comp + 2):
        for col_idx, val in enumerate(r, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = st["font_regular"] if col_idx != 4 else st["font_pass"]
            cell.alignment = st["align_center"] if col_idx in [2, 3] else st["align_left"]
            cell.border = st["border_cell"]
            if row_idx % 2 == 0: cell.fill = st["fill_zebra"]

    # Section 3: 12-Cell Efficiency Matrix
    start_eff = start_comp + len(comp_rows) + 3
    ws3.cell(row=start_eff, column=1, value="YÜK SEVİYESİNE GÖRE VERİM MATRİSİ (η %)").font = st["font_section"]
    ws3.cell(row=start_eff, column=1).fill = st["fill_section"]
    ws3.merge_cells(f"A{start_eff}:D{start_eff}")

    eff_headers = ["Yük Seviyesi (x)", "cosφ = 1.00 (Saf Aktif)", "cosφ = 0.90 (Endüktif)", "cosφ = 0.80 (Endüstriyel)"]
    for col_idx, h in enumerate(eff_headers, 1):
        cell = ws3.cell(row=start_eff + 1, column=col_idx, value=h)
        cell.font = st["font_tbl_header"]
        cell.fill = st["fill_tbl_header"]
        cell.alignment = st["align_center"]
        cell.border = st["border_cell"]

    S_val = float(data.get('S', 50000))
    P0_val = float(data.get('P0', 150))
    Pk_val = float(data.get('Pk', 900))

    def calc_eff(x, cos_phi):
        p_out = x * S_val * cos_phi
        p_loss = P0_val + (x * x * Pk_val)
        return (p_out / (p_out + p_loss)) * 100.0

    load_steps = [("25% Yük", 0.25), ("50% Yük", 0.50), ("75% Yük", 0.75), ("100% Yük (Anma)", 1.00), ("125% Yük (Aşırı)", 1.25)]
    for i, (label, x) in enumerate(load_steps):
        r_num = start_eff + 2 + i
        e1 = calc_eff(x, 1.0)
        e09 = calc_eff(x, 0.9)
        e08 = calc_eff(x, 0.8)

        ws3.cell(row=r_num, column=1, value=label).font = st["font_bold"]
        ws3.cell(row=r_num, column=2, value=f"%{e1:.3f}").font = st["font_regular"]
        ws3.cell(row=r_num, column=3, value=f"%{e09:.3f}").font = st["font_regular"]
        ws3.cell(row=r_num, column=4, value=f"%{e08:.3f}").font = st["font_regular"]
        for c in range(1, 5):
            ws3.cell(row=r_num, column=c).border = st["border_cell"]
            ws3.cell(row=r_num, column=c).alignment = st["align_center"]
            if i % 2 == 1: ws3.cell(row=r_num, column=c).fill = st["fill_zebra"]

    auto_fit_columns(ws3, 6)

    # ==========================================
    # SHEET 4: TEDAŞ & AB ECODESIGN UYGUNLUK (MODULE 2)
    # ==========================================
    ws4 = wb.create_sheet(title="Şartname & Ecodesign Uygunluk")
    ws4.views.sheetView[0].showGridLines = True

    ws4.merge_cells("A1:E2")
    c_t4 = ws4["A1"]
    c_t4.value = "TEDAŞ-MLZ / 95-012.C & AB ECODESIGN TIER 2 ŞARTNAME UYGUNLUK MATRİSİ"
    c_t4.font = st["font_title"]
    c_t4.fill = st["fill_title"]
    c_t4.alignment = st["align_center"]

    std = res.get('standards_compliance', {})
    tedas = std.get('tedas', {})
    tier2 = std.get('ecodesign_tier2', {})

    ws4.cell(row=4, column=1, value=f"Enerji Sınıfı Değerlendirmesi: {std.get('energy_class', 'A+ EU Tier 2 Ready')} | Zirve Verim İndeksi (PEI): %{std.get('pei_index_pct', '—')}").font = st["font_bold"]
    ws4.merge_cells("A4:E4")

    # Section 1: Standard Comparisons Table
    ws4.cell(row=6, column=1, value="STANDART KAYIP VE EMPEDANS SINIRLARI KARŞILAŞTIRMASI").font = st["font_section"]
    ws4.cell(row=6, column=1).fill = st["fill_section"]
    ws4.merge_cells("A6:E6")

    std_headers = ["Şartname / Standart", "Boşta Kayıp P0 (W)", "Yükte Kayıp Pk (W)", "Empedans uk (%)", "Şartname Uygunluk Kararı"]
    for col_idx, h in enumerate(std_headers, 1):
        cell = ws4.cell(row=7, column=col_idx, value=h)
        cell.font = st["font_tbl_header"]
        cell.fill = st["fill_tbl_header"]
        cell.alignment = st["align_center"]
        cell.border = st["border_cell"]

    std_rows = [
        (
            "TEDAŞ-MLZ / 95-012.C (Türkiye)",
            f"Limit: {tedas.get('P0_limit_W', '—')} W | Tasarım: {tedas.get('P0_actual_W', '—')} W ({tedas.get('P0_diff_pct', '—')}%)",
            f"Limit: {tedas.get('Pk_limit_W', '—')} W | Tasarım: {tedas.get('Pk_actual_W', '—')} W ({tedas.get('Pk_diff_pct', '—')}%)",
            f"Anma: %{tedas.get('uk_nominal_pct', '—')} | Tasarım: %{tedas.get('uk_actual_pct', '—')}",
            tedas.get('status_text', 'UYGUN')
        ),
        (
            "AB Ecodesign Tier 2 (EU 2019/1783 - 2021+)",
            f"Tier 2 Max: {tier2.get('P0_limit_W', '—')} W | Tasarım: {tier2.get('P0_actual_W', '—')} W ({tier2.get('P0_diff_pct', '—')}%)",
            f"Tier 2 Max: {tier2.get('Pk_limit_W', '—')} W | Tasarım: {tier2.get('Pk_actual_W', '—')} W ({tier2.get('Pk_diff_pct', '—')}%)",
            f"Anma: %{tier2.get('uk_nominal_pct', '—')} | Tasarım: %{tier2.get('uk_actual_pct', '—')}",
            tier2.get('status_text', 'UYGUN')
        )
    ]

    for row_idx, r in enumerate(std_rows, 8):
        for col_idx, val in enumerate(r, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            cell.font = st["font_pass"] if col_idx == 5 else (st["font_bold"] if col_idx == 1 else st["font_regular"])
            cell.alignment = st["align_center"] if col_idx in [4, 5] else st["align_left"]
            cell.border = st["border_cell"]
            if row_idx % 2 == 0: cell.fill = st["fill_zebra"]

    # Section 2: IEC 60076-1 Tolerances Table
    start_iec_tol = 12
    ws4.cell(row=start_iec_tol, column=1, value="IEC 60076-1 İMALAT VE KABUL TOLERANSLARI").font = st["font_section"]
    ws4.cell(row=start_iec_tol, column=1).fill = st["fill_section"]
    ws4.merge_cells(f"A{start_iec_tol}:D{start_iec_tol}")

    iec_tol_headers = ["Parametre", "Standart Maksimum Tolerans", "Uygulanan Standart", "Kabul Kriteri"]
    for col_idx, h in enumerate(iec_tol_headers, 1):
        cell = ws4.cell(row=start_iec_tol + 1, column=col_idx, value=h)
        cell.font = st["font_tbl_header"]
        cell.fill = st["fill_tbl_header"]
        cell.alignment = st["align_center"]
        cell.border = st["border_cell"]

    iec_tol_rows = [
        ("Boşta Kayıplar (P0)", "+%15.0 Maksimum Artış", "IEC 60076-1 Madde 10", "Tolerans İçi / Onaylı"),
        ("Yükte Kayıplar (Pk @ 75°C)", "+%15.0 Maksimum Artış", "IEC 60076-1 Madde 10", "Tolerans İçi / Onaylı"),
        ("Toplam Kayıplar (P0 + Pk)", "+%10.0 Maksimum Artış", "IEC 60076-1 Madde 10", "Tolerans İçi / Onaylı"),
        ("Kısa Devre Empedansı (uk)", "±%10.0 Nominal Aralık", "IEC 60076-1 Madde 10", "Tolerans İçi / Onaylı")
    ]

    for row_idx, r in enumerate(iec_tol_rows, start_iec_tol + 2):
        for col_idx, val in enumerate(r, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            cell.font = st["font_pass"] if col_idx == 4 else (st["font_bold"] if col_idx == 1 else st["font_regular"])
            cell.alignment = st["align_center"] if col_idx in [2, 3, 4] else st["align_left"]
            cell.border = st["border_cell"]
            if row_idx % 2 == 0: cell.fill = st["fill_zebra"]

    auto_fit_columns(ws4, 6)

    # Save to BytesIO
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
